#!/usr/bin/env python3
"""
fetch_ice_cot.py — Henter ICE Futures Europe COT-data (Brent, Gasoil, TTF Gas)

ICE publiserer ukentlig (fredag 18:30 London) rapport for tirsdag-posisjoner.
Dekker: Brent Crude, Low Sulphur Gas Oil (Gasoil), TTF Natural Gas

Output: data/ice_cot/latest.json  — samme struktur som fetch_all.py forventer
        data/ice_cot/history.json — rullende 26-ukers historikk

Brukes av fetch_all.py: Brent bruker ICE som primær COT-kilde (ICE er hjemmebørs
for Brent), med CFTC "Crude Oil, Light Sweet" som fallback.
"""

import json
import io
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl

BASE    = Path(__file__).parent
OUT_DIR = BASE / "data" / "ice_cot"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT     = OUT_DIR / "latest.json"
HIST    = OUT_DIR / "history.json"

# ICE-markeder vi bryr oss om (lowercase søkenøkkel → display-navn)
ICE_MARKETS = {
    "brent crude":        "ice brent crude",
    "low sulphur gasoil": "ice gasoil",
    "ttf natural gas":    "ice ttf gas",
    # Varianter brukt i ulike rapport-versjoner
    "brent":              "ice brent crude",
    "gasoil":             "ice gasoil",
    "ttf":                "ice ttf gas",
    "natural gas":        "ice ttf gas",
}

# ICE COT CSV-URLer — én fil per år (oppdateres ukentlig)
from datetime import datetime as _dt
_THIS_YEAR = _dt.now().year
ICE_URLS = [
    f"https://www.ice.com/publicdocs/futures/COTHist{_THIS_YEAR}.csv",
    f"https://www.ice.com/publicdocs/futures/COTHist{_THIS_YEAR - 1}.csv",
]
ICE_REPORT_PAGE = "https://www.ice.com/report/122"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/vnd.ms-excel,*/*",
    "Referer": "https://www.ice.com/",
}


# ── Nedlasting ───────────────────────────────────────────────────────────────

def find_excel_url():
    """Prøv å finne direkte Excel-URL fra rapport-siden."""
    try:
        r = requests.get(ICE_REPORT_PAGE, headers=HEADERS, timeout=15)
        # Let etter .xlsx eller .xls lenker i HTML
        urls = re.findall(r'href=["\']([^"\']*\.xlsx?)["\']', r.text, re.I)
        for u in urls:
            if "cot" in u.lower() or "commitments" in u.lower() or "futures" in u.lower():
                if u.startswith("http"):
                    return u
                return "https://www.ice.com" + u
    except Exception as e:
        print(f"  Kunne ikke hente rapport-side: {e}")
    return None


def download_csv():
    """Last ned ICE COT CSV. Returnerer tekst eller None."""
    for url in ICE_URLS:
        try:
            print(f"  Prøver: {url}")
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            if len(r.content) < 1000:
                continue
            if b"<!DOCTYPE" in r.content[:200]:
                continue
            print(f"    OK — {len(r.content)//1024} KB")
            return r.content.decode("utf-8", errors="replace")
        except Exception as e:
            print(f"    FEIL: {e}")
    return None


# ── Excel-parsing ────────────────────────────────────────────────────────────

def match_market(name):
    """Returner canonical market-nøkkel hvis name matcher ICE_MARKETS."""
    n = str(name).lower()
    for search_key, canonical in ICE_MARKETS.items():
        if search_key in n:
            return canonical
    return None


def normalize(s):
    """Normaliser celleinnhold til lowercase streng."""
    if s is None:
        return ""
    return str(s).strip().lower()


def parse_ice_csv(text):
    """
    Parse ICE COT CSV (CFTC-format).
    Henter siste rad per marked (nyeste rapport øverst eller nederst).
    Returnerer dict: {market_key → row_data}
    """
    import csv, io as _io
    results = {}
    reader = csv.DictReader(_io.StringIO(text))
    rows_by_market = {}

    for row in reader:
        name = row.get("Market_and_Exchange_Names", "") or row.get("market_and_exchange_names", "")
        if not name:
            # Prøv første kolonne uansett navn
            name = next(iter(row.values()), "")

        mkey = match_market(name)
        if mkey is None:
            continue

        # Finn kolonnene (CFTC-format)
        def g(keys):
            for k in keys:
                for col, val in row.items():
                    if k.lower() in col.lower():
                        return val
            return "0"

        mm_long  = safe_num(g(["NonComm_Positions_Long", "Noncommercial_Long", "Managed_Money_Long", "Money_Manager_Long"]))
        mm_short = safe_num(g(["NonComm_Positions_Short", "Noncommercial_Short", "Managed_Money_Short", "Money_Manager_Short"]))
        oi       = safe_num(g(["Open_Interest", "open_interest"]))
        chg_long = safe_num(g(["Change_in_NonComm_Long", "Change_NonComm_Long"]))
        chg_short= safe_num(g(["Change_in_NonComm_Short", "Change_NonComm_Short"]))

        date_str = row.get("Report_Date_as_YYYY-MM-DD", "") or row.get("As_of_Date_In_Form_YYMMDD", "")

        # Behold siste (nyeste) rad per marked
        existing = rows_by_market.get(mkey)
        if existing is None or date_str >= existing.get("date", ""):
            rows_by_market[mkey] = {
                "mm_long": mm_long, "mm_short": mm_short,
                "mm_net": mm_long - mm_short,
                "oi": oi, "chg": chg_long - chg_short,
                "date": date_str,
            }

    for mkey, d in rows_by_market.items():
        results[mkey] = d
    return results


def safe_num(s):
    try:
        return int(float(str(s).replace(",", "").strip()))
    except Exception:
        return 0


def parse_ice_excel(data):
    """
    Parse ICE COT Excel. Returnerer dict: {market_key → row_data}
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Finn header-rad (inneholder "long" og "short")
        header_row = None
        header_idx = None
        for i, row in enumerate(rows[:20]):
            row_str = " ".join(normalize(c) for c in row if c is not None)
            if "long" in row_str and "short" in row_str:
                header_row = [normalize(c) for c in row]
                header_idx = i
                break

        if header_row is None:
            # Prøv Format B: let etter "managed money" direkte
            results.update(_parse_format_b(rows))
            continue

        # Format A: parse med header
        results.update(_parse_format_a(rows, header_row, header_idx))

    return results


def _parse_format_a(rows, header_row, header_idx):
    """
    Format A: Bred tabell med én rad per marked.
    Kolonner inkluderer market-name, date, og per-kategori long/short/net.
    """
    results = {}
    h = header_row

    # Finn kolonne-indekser
    def col(keywords):
        for kw in keywords:
            for i, c in enumerate(h):
                if kw in c:
                    return i
        return None

    # Prøv å finne relevante kolonner
    # ICE bruker typisk: "market", "date", "open interest"
    # og per kategori: "managed money long", "managed money short"
    market_col  = col(["market", "commodity", "contract"])
    date_col    = col(["date", "as of"])
    oi_col      = col(["open interest", "total open"])

    # Managed Money kolonner (= spekulanter, tilsvarer CFTC Managed Money)
    mm_long_col  = None
    mm_short_col = None
    for i, c in enumerate(h):
        if "managed" in c and "long" in c:
            mm_long_col = i
        if "managed" in c and "short" in c:
            mm_short_col = i

    # Fallback: "non-commercial" = spekulanter i eldre format
    if mm_long_col is None:
        for i, c in enumerate(h):
            if ("non-commercial" in c or "noncommercial" in c) and "long" in c:
                mm_long_col = i
            if ("non-commercial" in c or "noncommercial" in c) and "short" in c:
                mm_short_col = i

    if market_col is None or mm_long_col is None:
        return results

    # Parse datarader
    for row in rows[header_idx + 1:]:
        if not row or row[market_col] is None:
            continue
        market_name = normalize(row[market_col])
        if not market_name:
            continue

        # Matcher mot ICE_MARKETS
        matched_key = None
        for search_key, canonical in ICE_MARKETS.items():
            if search_key in market_name:
                matched_key = canonical
                break
        if matched_key is None:
            continue

        def safe_int(col_i):
            if col_i is None or col_i >= len(row):
                return 0
            v = row[col_i]
            if v is None:
                return 0
            try:
                return int(float(str(v).replace(",", "").strip()))
            except Exception:
                return 0

        mm_long  = safe_int(mm_long_col)
        mm_short = safe_int(mm_short_col) if mm_short_col else 0
        mm_net   = mm_long - mm_short
        oi       = safe_int(oi_col) if oi_col else max(abs(mm_net) * 5, 1)

        raw_date = row[date_col] if date_col and date_col < len(row) else None
        date_str = _parse_date(raw_date)

        if mm_long == 0 and mm_short == 0:
            continue

        results[matched_key] = {
            "mm_long":  mm_long,
            "mm_short": mm_short,
            "mm_net":   mm_net,
            "oi":       oi,
            "date":     date_str,
        }
    return results


def _parse_format_b(rows):
    """
    Format B: Én rad per marked×kategori. Typisk kolonner:
    Date | Market | Category | Long | Short | Change in Long | Change in Short | % OI | Traders
    """
    results = {}
    # Finn header-rad med "category" eller "contract"
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:30]):
        row_str = " ".join(normalize(c) for c in row if c is not None)
        if ("category" in row_str or "contract" in row_str) and "long" in row_str:
            header_row = [normalize(c) for c in row]
            header_idx = i
            break

    if header_row is None:
        return results

    h = header_row
    def col(keywords):
        for kw in keywords:
            for i, c in enumerate(h):
                if kw in c:
                    return i
        return None

    date_col     = col(["date"])
    market_col   = col(["market", "contract", "commodity"])
    category_col = col(["category", "trader type", "participant"])
    long_col     = col(["long"])
    short_col    = col(["short"])
    oi_col       = col(["open interest"])

    if market_col is None or long_col is None:
        return results

    # Samle rader per marked
    by_market = {}
    for row in rows[header_idx + 1:]:
        if not row or (market_col and row[market_col] is None):
            continue
        market_name = normalize(row[market_col]) if market_col < len(row) else ""
        matched_key = None
        for search_key, canonical in ICE_MARKETS.items():
            if search_key in market_name:
                matched_key = canonical
                break
        if matched_key is None:
            continue

        def safe_int(col_i):
            if col_i is None or col_i >= len(row):
                return 0
            v = row[col_i]
            if v is None:
                return 0
            try:
                return int(float(str(v).replace(",", "").strip()))
            except Exception:
                return 0

        category = normalize(row[category_col]) if category_col and category_col < len(row) else ""
        is_spec = any(k in category for k in ["managed", "non-commercial", "noncommercial",
                                                "leveraged", "hedge fund"])
        if not is_spec:
            continue

        raw_date = row[date_col] if date_col and date_col < len(row) else None
        date_str = _parse_date(raw_date)

        key_entry = by_market.setdefault(matched_key, {
            "mm_long": 0, "mm_short": 0, "mm_net": 0, "oi": 0, "date": date_str
        })
        key_entry["mm_long"]  += safe_int(long_col)
        key_entry["mm_short"] += safe_int(short_col)
        key_entry["mm_net"]    = key_entry["mm_long"] - key_entry["mm_short"]
        if oi_col:
            oi = safe_int(oi_col)
            if oi > key_entry["oi"]:
                key_entry["oi"] = oi
        if date_str and not key_entry["date"]:
            key_entry["date"] = date_str

    results.update({k: v for k, v in by_market.items() if v["mm_long"] or v["mm_short"]})
    return results


def _parse_date(raw):
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%B %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s[:10] if len(s) >= 10 else s


# ── Historikk og deltaberegning ──────────────────────────────────────────────

def load_history():
    if HIST.exists():
        try:
            return json.loads(HIST.read_text())
        except Exception:
            pass
    return {}


def update_history(history, parsed):
    """Legg til nye ukentlige verdier i historikk (maks 26 uker = 6 mnd)."""
    for key, data in parsed.items():
        entries = history.setdefault(key, [])
        date = data.get("date", "")
        # Ikke dupliser samme dato
        if entries and entries[-1].get("date") == date:
            continue
        entries.append({
            "date":   date,
            "net":    data["mm_net"],
            "long":   data["mm_long"],
            "short":  data["mm_short"],
        })
        history[key] = entries[-26:]   # Behold 26 siste uker
    return history


def calc_change(history, key, current_net):
    """Beregn ukentlig endring fra forrige rapporterte verdi."""
    entries = history.get(key, [])
    if len(entries) >= 2:
        return current_net - entries[-2]["net"]
    return 0


# ── Bygg output-struktur ─────────────────────────────────────────────────────

def build_output(parsed, history):
    """
    Bygg output som matcher strukturen fetch_all.py forventer fra combined/latest.json.
    Spesielt: spekulanter.net, open_interest, change_spec_net, spec_net_history, date, report
    """
    results = []
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for canonical_key, data in parsed.items():
        net    = data["mm_net"]
        oi     = data.get("oi") or max(abs(net) * 8, 1)
        date   = data.get("date") or today
        change = calc_change(history, canonical_key, net)
        hist   = [e["net"] for e in history.get(canonical_key, [])]

        results.append({
            "market":           canonical_key,
            "display_name":     canonical_key.replace("ice ", "ICE ").title(),
            "spekulanter": {
                "long":  data["mm_long"],
                "short": data["mm_short"],
                "net":   net,
                "label": "Managed Money",
            },
            "open_interest":    oi,
            "change_spec_net":  change,
            "spec_net_history": hist,
            "date":             date,
            "report":           "ice",
        })

    return results


# ── Hovedkjøring ──────────────────────────────────────────────────────────────

def main():
    print("Henter ICE Futures Europe COT-data...")

    raw = download_csv()
    if raw is None:
        print("  FEIL: Kunne ikke laste ned ICE COT CSV — ingen data skrevet")
        return False

    try:
        parsed = parse_ice_csv(raw)
    except Exception as e:
        print(f"  FEIL ved parsing av CSV: {e}")
        return False

    if not parsed:
        print("  FEIL: Ingen ICE-markeder funnet i Excel-filen")
        print("  (ICE har muligens endret filformat)")
        return False

    history = load_history()
    history = update_history(history, parsed)
    HIST.write_text(json.dumps(history, ensure_ascii=False, indent=2))

    output_list = build_output(parsed, history)

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    output = {
        "generated": now_str,
        "source":    "ICE Futures Europe",
        "markets":   output_list,
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2))

    print(f"  OK → {len(output_list)} markeder lagret")
    for m in output_list:
        net = m["spekulanter"]["net"]
        pct = net / m["open_interest"] * 100 if m["open_interest"] else 0
        print(f"    {m['display_name']:30} net={net:+,}  ({pct:+.1f}% OI)  {m['date']}")

    return True


if __name__ == "__main__":
    ok = main()
    raise SystemExit(0 if ok else 1)
