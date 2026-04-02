#!/usr/bin/env python3
"""
fetch_euronext_cot.py — Henter Euronext COT-data for landbruksråvarer (MiFID II)

Euronext publiserer hver onsdag ettermiddag, posisjoner per foregående fredag.
Dekker: Milling Wheat (EBM), Rapeseed (EBR), Corn/Maize (EMA)

Rapporten (Excel, "agri-commodities") viser per kategori:
  antall personer, long/short, endring, % av open interest

MiFID II-kategorier (brukt av Euronext):
  - Investment Funds          → tilsvarer CFTC Managed Money (spekulanter)
  - Investment Firms/Banks    → Swap Dealers / mellommenn
  - Commercial Undertakings   → Producers/Commercials (hedgere)
  - Other Financial           → Other Reportables

Output: data/euronext_cot/latest.json  — samme struktur som fetch_agri.py forventer
        data/euronext_cot/history.json — rullende 26-ukers historikk
"""

import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl

BASE    = Path(__file__).parent
OUT_DIR = BASE / "data" / "euronext_cot"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT  = OUT_DIR / "latest.json"
HIST = OUT_DIR / "history.json"

# Euronext-markeder → intern crop-nøkkel
# Søker case-insensitivt i markedsnavn/produktnavn
EURONEXT_MARKETS = {
    # Milling Wheat
    "milling wheat":   "wheat",
    "wheat milling":   "wheat",
    "ebm":             "wheat",
    "farine":          "wheat",
    "blé":             "wheat",
    "ble":             "wheat",
    # Rapeseed / Colza
    "rapeseed":        "canola",
    "colza":           "canola",
    "ebr":             "canola",
    "canola":          "canola",
    # Corn / Maize
    "corn":            "corn",
    "maize":           "corn",
    "ema":             "corn",
    "mais":            "corn",
    # Sunflower (bonus — vises hvis det finnes)
    "sunflower":       "sunflower",
    "tournesol":       "sunflower",
}

# MiFID II-kategorier → type (viktigst: investment_funds = spekulanter)
SPEC_KEYWORDS    = ["investment fund", "fonds d'investissement", "managed", "fund"]
COMMERCIAL_KEYWORDS = ["commercial undertaking", "entreprise commerciale", "commercial"]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://live.euronext.com/",
}
HEADERS_XLSX = {**HEADERS,
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/vnd.ms-excel,*/*",
}

from datetime import date as _date
import itertools

def _candidate_urls():
    """Generer kandidat-URLer basert på dato (siste 8 uker)."""
    today = _date.today()
    urls = []
    # live.euronext.com — prøv siste 8 uker med dato-prefix
    for weeks_ago in range(0, 8):
        from datetime import timedelta
        d = today - timedelta(weeks=weeks_ago)
        month_str = d.strftime("%Y-%m")
        urls.append(
            f"https://live.euronext.com/sites/default/files/{month_str}/agri-commodities-cot.xlsx"
        )
    # Fallback uten dato-prefix
    urls += [
        "https://live.euronext.com/sites/default/files/agri-commodities-cot.xlsx",
        "https://live.euronext.com/sites/default/files/agri-cot.xlsx",
    ]
    return urls

# Euronext COT rapport-sider (skrapes for Excel-lenke)
EURONEXT_PAGES = [
    "https://live.euronext.com/en/products/commodities/commitments_of_traders",
    "https://www.euronext.com/en/markets/derivatives/commitment-of-traders",
]


# ── Nedlasting ───────────────────────────────────────────────────────────────

def find_excel_url():
    """Skrap Euronext COT-side for Excel-lenke med 'agri' i filnavnet."""
    for page_url in EURONEXT_PAGES:
        try:
            r = requests.get(page_url, headers=HEADERS, timeout=20)
            urls = re.findall(r'href=["\']([^"\']*agri[^"\']*\.xlsx?)["\']', r.text, re.I)
            if not urls:
                urls = re.findall(r'href=["\']([^"\']*\.xlsx?)["\']', r.text, re.I)
            for u in urls:
                if not u.startswith("http"):
                    base = "https://live.euronext.com" if "live.euronext" in page_url else "https://www.euronext.com"
                    u = base + u
                return u
        except Exception as e:
            print(f"  Kunne ikke hente side {page_url}: {e}")
    return None


def download_with_playwright():
    """Bruk Playwright (headless Chrome) for å laste ned Excel via ekte nettleserøkt."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()
        try:
            # Fang opp nettverksforespørsler for å finne xlsx-URL
            captured_xlsx = []
            def on_request(req):
                url = req.url
                if ".xls" in url.lower() or "agri" in url.lower() and "cot" in url.lower():
                    captured_xlsx.append(url)
            def on_response(resp):
                url = resp.url
                ct = resp.headers.get("content-type", "")
                if "spreadsheet" in ct or "excel" in ct or ".xls" in url.lower():
                    captured_xlsx.append(url)
            page.on("request", on_request)
            page.on("response", on_response)

            print("  Playwright: åpner Euronext COT-side...")
            page.goto(EURONEXT_PAGES[0], wait_until="networkidle", timeout=45000)
            page.wait_for_timeout(3000)

            # Sjekk også HTML-kilde for xlsx-referanser
            html = page.content()
            xlsx_in_html = re.findall(r'["\']([^"\']*agri[^"\']*\.xlsx?)["\']', html, re.I)
            xlsx_in_html += re.findall(r'["\']([^"\']*cot[^"\']*\.xlsx?)["\']', html, re.I)
            print(f"  xlsx i HTML: {xlsx_in_html[:3]}")
            print(f"  xlsx fra nettverksforespørsler: {captured_xlsx[:3]}")

            # Finn URL fra nettverksforespørsler/HTML
            raw_url = (captured_xlsx + xlsx_in_html + [None])[0]
            xlsx_url = None
            if raw_url:
                xlsx_url = raw_url if raw_url.startswith("http") else "https://live.euronext.com" + raw_url
                print(f"  Fant xlsx-URL: {xlsx_url}")

            # Prøv å klikke nedlastingsknapper og fange responsen
            if not xlsx_url:
                dl_response = [None]
                def on_resp_body(resp):
                    ct = resp.headers.get("content-type", "")
                    if "spreadsheet" in ct or "excel" in ct or (resp.status == 200 and ".xls" in resp.url.lower()):
                        try:
                            dl_response[0] = resp.body()
                        except Exception:
                            pass
                page.on("response", on_resp_body)

                # Let etter knapper/lenker med "download", "excel", "cot", "agri" tekst
                selectors = [
                    "a[href*='xlsx']", "a[href*='xls']",
                    "button:has-text('Download')", "button:has-text('Excel')",
                    "a:has-text('Download')", "a:has-text('Excel')",
                    "a:has-text('COT')", "a:has-text('agri')",
                    "[data-file*='xlsx']", "[data-url*='xlsx']",
                ]
                for sel in selectors:
                    try:
                        el = page.query_selector(sel)
                        if el:
                            href = el.get_attribute("href") or el.get_attribute("data-url") or ""
                            print(f"  Fant element ({sel}): {href[:80]}")
                            with page.expect_download(timeout=8000) as dl_info:
                                el.click()
                            dl = dl_info.value
                            data = Path(dl.path()).read_bytes()
                            if len(data) > 500:
                                print(f"    OK via klikk — {len(data)//1024} KB")
                                browser.close()
                                return data
                    except Exception:
                        if dl_response[0] and len(dl_response[0]) > 500:
                            print(f"    OK via respons-fang — {len(dl_response[0])//1024} KB")
                            browser.close()
                            return dl_response[0]

            candidates = ([xlsx_url] if xlsx_url else []) + _candidate_urls()

            # Bruk Playwright API-context (har nettleserens cookies/session)
            api_ctx = ctx.request
            for url in candidates:
                if not url:
                    continue
                try:
                    print(f"  Playwright API: {url}")
                    resp = api_ctx.get(url, headers={"Accept": "*/*"}, timeout=20000)
                    if resp.ok:
                        data = resp.body()
                        if len(data) > 500 and not data[:5].startswith(b"<!DOC"):
                            print(f"    OK — {len(data)//1024} KB")
                            browser.close()
                            return data
                    else:
                        print(f"    HTTP {resp.status}")
                except Exception as e:
                    print(f"    FEIL: {e}")
        except Exception as e:
            print(f"  Playwright FEIL: {e}")
        finally:
            browser.close()
    return None


def download_excel():
    """Last ned Euronext agri COT Excel. Returnerer bytes eller None."""
    # Prøv Playwright først (ekte nettleser, omgår Cloudflare)
    data = download_with_playwright()
    if data:
        return data

    # Fallback: vanlig requests med session
    print("  Playwright ikke tilgjengelig/feilet — prøver requests...")
    sess = requests.Session()
    for page in EURONEXT_PAGES:
        try:
            sess.get(page, headers=HEADERS, timeout=15)
            break
        except Exception:
            continue

    for url in _candidate_urls():
        try:
            print(f"  Prøver: {url}")
            r = sess.get(url, headers=HEADERS_XLSX, timeout=30, allow_redirects=True)
            r.raise_for_status()
            ctype = r.headers.get("Content-Type", "")
            if len(r.content) < 500:
                continue
            if "html" in ctype.lower() and b"<!DOCTYPE" in r.content[:200]:
                continue
            print(f"    OK — {len(r.content)//1024} KB")
            return r.content
        except Exception as e:
            print(f"    FEIL: {e}")
    return None


# ── Excel-parsing ────────────────────────────────────────────────────────────

def normalize(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def safe_int(v):
    if v is None:
        return 0
    try:
        return int(float(str(v).replace(",", "").replace(" ", "").strip()))
    except Exception:
        return 0


def match_market(cell_val):
    """Returner crop_key hvis celleverdien matcher et kjent Euronext-marked."""
    s = normalize(cell_val)
    for keyword, crop in EURONEXT_MARKETS.items():
        if keyword in s:
            return crop
    return None


def is_speculator_row(cell_val):
    s = normalize(cell_val)
    return any(k in s for k in SPEC_KEYWORDS)


def parse_euronext_excel(data):
    """
    Parser Euronext agri COT Excel.

    Euronext-formater vi støtter:
    Format A — Bred tabell: én rad per marked, kolonner per kategori×long/short
    Format B — Lang tabell: én rad per marked×kategori
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws    = wb[sheet_name]
        rows  = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Søk etter header-rad som inneholder "long" og "short"
        header_idx  = None
        header_row  = None
        for i, row in enumerate(rows[:30]):
            s = " ".join(normalize(c) for c in row if c is not None)
            if "long" in s and "short" in s:
                header_row = [normalize(c) for c in row]
                header_idx = i
                break

        if header_row:
            r = _parse_format_a(rows, header_row, header_idx)
            if r:
                results.update(r)
                continue

        # Fallback: Format B (lang tabell)
        r = _parse_format_b(rows)
        if r:
            results.update(r)

    return results


def _parse_format_a(rows, header, header_idx):
    """
    Bred tabell: én rad per marked.
    Kolonner inneholder markedsnavn, dato, og verdier per kategori.
    Vi leter spesielt etter "investment fund" long/short-kolonner.
    """
    results = {}
    h = header

    def find_col(keywords):
        """Finn kolonne-indeks som inneholder alle keywords i overskriften (evt. kombinert med naborade)."""
        for i, c in enumerate(h):
            if all(kw in c for kw in keywords):
                return i
        return None

    def find_col_any(keywords):
        for i, c in enumerate(h):
            if any(kw in c for kw in keywords):
                return i
        return None

    market_col  = find_col_any(["product", "market", "contract", "commodity", "instrument"])
    date_col    = find_col_any(["date", "reference", "reporting"])
    oi_col      = find_col_any(["open interest", "open_interest", "total oi"])

    # Investment Funds (spekulanter) — lang/short
    spec_long_col  = None
    spec_short_col = None
    for i, c in enumerate(h):
        if any(k in c for k in SPEC_KEYWORDS):
            if "long" in c and spec_long_col is None:
                spec_long_col = i
            elif "short" in c and spec_short_col is None:
                spec_short_col = i

    # Fallback: "non-commercial" (eldre format)
    if spec_long_col is None:
        for i, c in enumerate(h):
            if ("non-commercial" in c or "noncommercial" in c):
                if "long" in c and spec_long_col is None:
                    spec_long_col = i
                elif "short" in c and spec_short_col is None:
                    spec_short_col = i

    if market_col is None or spec_long_col is None:
        return {}

    # Parse datarader
    for row in rows[header_idx + 1:]:
        if not row or row[market_col] is None:
            continue
        crop = match_market(row[market_col])
        if crop is None:
            continue

        mm_long  = safe_int(row[spec_long_col]  if spec_long_col  < len(row) else None)
        mm_short = safe_int(row[spec_short_col] if spec_short_col and spec_short_col < len(row) else None)
        mm_net   = mm_long - mm_short
        oi       = safe_int(row[oi_col] if oi_col and oi_col < len(row) else None)
        if oi == 0:
            oi = max(abs(mm_net) * 6, 1)

        raw_date = row[date_col] if date_col and date_col < len(row) else None
        date_str = _parse_date(raw_date)

        if mm_long == 0 and mm_short == 0:
            continue

        # Ikke overskriv med dårligere data
        if crop in results and results[crop]["mm_long"] > mm_long and mm_long == 0:
            continue

        results[crop] = {
            "mm_long":  mm_long,
            "mm_short": mm_short,
            "mm_net":   mm_net,
            "oi":       oi,
            "date":     date_str,
        }

    return results


def _parse_format_b(rows):
    """
    Lang tabell: én rad per marked × kategori.
    Typiske kolonner: Product | Category | Date | Long | Short | Change Long | Change Short | % OI
    """
    # Finn header
    header_idx = None
    header_row = None
    for i, row in enumerate(rows[:30]):
        s = " ".join(normalize(c) for c in row if c is not None)
        if ("category" in s or "participant" in s or "type" in s) and "long" in s:
            header_row = [normalize(c) for c in row]
            header_idx = i
            break

    if header_row is None:
        return {}

    h = header_row
    def find_col_any(kws):
        for i, c in enumerate(h):
            if any(kw in c for kw in kws):
                return i
        return None

    market_col   = find_col_any(["product", "market", "contract", "commodity"])
    category_col = find_col_any(["category", "participant", "type", "trader"])
    date_col     = find_col_any(["date", "reference"])
    long_col     = find_col_any(["long"])
    short_col    = find_col_any(["short"])
    oi_col       = find_col_any(["open interest", "oi"])

    if market_col is None or long_col is None:
        return {}

    by_crop = {}
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        crop = match_market(row[market_col] if market_col < len(row) else None)
        if crop is None:
            continue

        cat_val = row[category_col] if category_col and category_col < len(row) else ""
        if not is_speculator_row(cat_val):
            continue

        mm_long  = safe_int(row[long_col]  if long_col  < len(row) else None)
        mm_short = safe_int(row[short_col] if short_col and short_col < len(row) else None)
        oi       = safe_int(row[oi_col] if oi_col and oi_col < len(row) else None)
        raw_date = row[date_col] if date_col and date_col < len(row) else None
        date_str = _parse_date(raw_date)

        if mm_long == 0 and mm_short == 0:
            continue

        entry = by_crop.setdefault(crop, {"mm_long": 0, "mm_short": 0, "mm_net": 0,
                                           "oi": 0, "date": date_str})
        entry["mm_long"]  += mm_long
        entry["mm_short"] += mm_short
        entry["mm_net"]    = entry["mm_long"] - entry["mm_short"]
        if oi > entry["oi"]:
            entry["oi"] = oi

    return by_crop


def _parse_date(raw):
    if raw is None:
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                "%d.%m.%Y", "%B %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s[:10] if len(s) >= 10 else s


# ── Historikk ────────────────────────────────────────────────────────────────

def load_history():
    if HIST.exists():
        try:
            return json.loads(HIST.read_text())
        except Exception:
            pass
    return {}


def update_history(history, parsed):
    for crop, data in parsed.items():
        entries = history.setdefault(crop, [])
        date = data.get("date", "")
        if entries and entries[-1].get("date") == date:
            continue
        entries.append({"date": date, "net": data["mm_net"],
                        "long": data["mm_long"], "short": data["mm_short"]})
        history[crop] = entries[-26:]
    return history


def calc_change(history, crop, current_net):
    entries = history.get(crop, [])
    if len(entries) >= 2:
        return current_net - entries[-2]["net"]
    return 0


# ── Bygg output ───────────────────────────────────────────────────────────────

CROP_DISPLAY = {
    "wheat":     "Milling Wheat (Euronext EBM)",
    "canola":    "Rapeseed (Euronext EBR)",
    "corn":      "Corn/Maize (Euronext EMA)",
    "sunflower": "Sunflower Seed (Euronext)",
}


def build_output(parsed, history):
    results = {}
    today   = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for crop, data in parsed.items():
        net    = data["mm_net"]
        oi     = data.get("oi") or max(abs(net) * 7, 1)
        date   = data.get("date") or today
        change = calc_change(history, crop, net)
        hist   = [e["net"] for e in history.get(crop, [])]

        results[crop] = {
            "crop":        crop,
            "display":     CROP_DISPLAY.get(crop, f"Euronext {crop.title()}"),
            "spekulanter": {
                "long":  data["mm_long"],
                "short": data["mm_short"],
                "net":   net,
                "label": "Investment Funds (MiFID II)",
            },
            "open_interest":   oi,
            "change_spec_net": change,
            "spec_net_history": hist,
            "date":   date,
            "report": "euronext",
            "source": "Euronext",
        }

    return results


# ── Hoved ────────────────────────────────────────────────────────────────────

def main():
    print("Henter Euronext agri COT-data (MiFID II)...")

    raw = download_excel()
    if raw is None:
        print("  FEIL: Kunne ikke laste ned Euronext COT Excel")
        return False

    try:
        parsed = parse_euronext_excel(raw)
    except Exception as e:
        print(f"  FEIL ved parsing: {e}")
        return False

    if not parsed:
        print("  FEIL: Ingen Euronext-markeder funnet i Excel-filen")
        return False

    history = load_history()
    history = update_history(history, parsed)
    HIST.write_text(json.dumps(history, ensure_ascii=False, indent=2))

    output_data = build_output(parsed, history)
    now_str     = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    out = {
        "generated": now_str,
        "source":    "Euronext Derivatives — MiFID II COT",
        "markets":   output_data,
    }
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2))

    print(f"  OK → {len(output_data)} avlinger lagret")
    for crop, d in output_data.items():
        net = d["spekulanter"]["net"]
        oi  = d["open_interest"]
        pct = net / oi * 100 if oi else 0
        print(f"    {d['display']:40} net={net:+,}  ({pct:+.1f}% OI)  {d['date']}")

    return True


if __name__ == "__main__":
    ok = main()
    raise SystemExit(0 if ok else 1)
